import pandas as pd
import openpyxl as xl
i = 1
read_file = pd.read_csv (r'C:\Users\48941.FRH179290\Desktop\SWIM2\Strat_' + str(i) + r'_output\combined' + str(i) + r'.csv')
read_file.to_excel (r'C:\Users\48941.FRH179290\Desktop\SWIM2\excelfolder\Strat' + str(i) + r'.xlsx', index = None, header=True)

wb = xl.load_workbook(r'C:\Users\48941.FRH179290\Desktop\SWIM2\excelfolder\Strat' + str(i) + r'.xlsx')
sheet = wb['Sheet1']
a = sheet.max_row
for column in range(2, sheet.max_column + 1):
    n = 0.0
    for row in range(2, a + 1):
        cell = sheet.cell(row, column)
        n = n + cell.value
    cell = sheet.cell(a + 1,column)
    cell.value = n
wb.save(r'C:\Users\48941.FRH179290\Desktop\SWIM2\excelfolder\Strat' + str(i) + r'.xlsx')


